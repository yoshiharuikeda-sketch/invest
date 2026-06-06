# -*- coding: utf-8 -*-
"""仮想売買（ペーパートレード）損益記録スクリプト

DRY検証期間用。実発注はせず、当日のシグナルに対して kabuステーションの板から
始値(OpeningPrice)・終値(CurrentPrice)を取得し、寄付き→引けの仮想損益を計算して
paper_trade_history.xlsx に1ファイルへ蓄積する。

  python -X utf8 paper_trade.py        # 当日分を計算して蓄積・通知

データソース優先:
  - kabuステーション /board の OpeningPrice(始値) / CurrentPrice(終値)
  - 価格が取れない銘柄はスキップ（備考に記録）

数量・スキップ条件は本番(kabu_order.py)と同じ:
  - 数量 = int(PORTFOLIO_VALUE * |ポジション| / 始値 / lot) * lot
  - SHORT非対応銘柄(SHORT_SKIP_TICKERS)のSHORTはスキップ（本番同様、約定不能のため）
  - 損益 = sign(ポジション) * 数量 * (終値 - 始値)
"""

import csv
import io
import os
import sys
from datetime import datetime
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np

import kabu_order as ko

DIR = Path(__file__).parent
PAPER_HIST_CSV    = DIR / "paper_pnl_history.csv"
PAPER_DETAIL_CSV  = DIR / "paper_pnl_detail_history.csv"
PAPER_XLSX        = DIR / "paper_trade_history.xlsx"

HIST_FIELDS   = ["日付", "合計損益(円)", "損益率(%)", "勝敗", "ポジション数"]
DETAIL_FIELDS = ["日付", "銘柄", "名称", "方向", "始値", "終値", "数量",
                 "損益(円)", "OC騰落率(%)", "備考"]


# DRY検証期間の仮想運用資金。本番の .env_windows PORTFOLIO_VALUE(30万) とは独立に
# ここで100万円に固定する（本番復帰時に誤って大きい金額で実発注しないための安全分離）。
VERIFY_CAPITAL = 1_000_000.0


def _latest_signal():
    files = sorted([f for f in os.listdir(DIR)
                    if f.startswith("signal_") and f.endswith(".csv")], reverse=True)
    if not files:
        return None, None
    fn = files[0]
    date_str = fn.replace("signal_", "").replace(".csv", "")  # 米国市場日付(YYYYMMDD)
    date_fmt = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    df = pd.read_csv(DIR / fn, encoding="utf-8-sig")
    return df, date_fmt


def _board_open_close(token: str, code: str):
    """kabuステーション板から (始値, 終値) を返す。取れなければ (None, None)。"""
    try:
        b = ko.get_board(token, code)
        op = b.get("OpeningPrice") or 0
        cl = b.get("CurrentPrice") or b.get("CalcPrice") or 0
        return (float(op) if op else None, float(cl) if cl else None)
    except Exception:
        return (None, None)


def _upsert_csv(path: Path, fields: list, rows: list, key_cols: list):
    """既存CSVを読み、key_cols が一致する行を除いてから rows を追記保存（重複日付の置換）。"""
    existing = []
    if path.exists():
        with open(path, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                existing.append(r)
    new_keys = {tuple(str(r[k]) for k in key_cols) for r in rows}
    kept = [r for r in existing if tuple(str(r.get(k, "")) for k in key_cols) not in new_keys]
    allrows = kept + rows
    # 日付でソート
    try:
        allrows.sort(key=lambda r: str(r.get("日付", "")))
    except Exception:
        pass
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(allrows)


def _save_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️ openpyxl 未インストール → Excel保存スキップ")
        return
    if not PAPER_HIST_CSV.exists():
        return
    df_hist = pd.read_csv(PAPER_HIST_CSV, encoding="utf-8-sig")
    df_det  = (pd.read_csv(PAPER_DETAIL_CSV, encoding="utf-8-sig")
               if PAPER_DETAIL_CSV.exists() else pd.DataFrame())

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6A1B9A")  # 紫＝仮想（本番1565C0と区別）
    green_font  = Font(color="2E7D32", bold=True)
    red_font    = Font(color="C62828", bold=True)
    gray_font   = Font(color="9E9E9E")
    center      = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin        = Side(border_style="thin", color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write(ws, df, money_cols, pct_cols):
        cols = list(df.columns)
        for ci, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
        for ri, row in enumerate(df.itertuples(index=False), start=2):
            for ci, (col, val) in enumerate(zip(cols, row), start=1):
                cell = ws.cell(row=ri, column=ci)
                if pd.isna(val) or val == "":
                    cell.value, cell.font, cell.alignment = "—", gray_font, center
                elif col in money_cols:
                    try:
                        v = float(val); cell.value = v
                        cell.number_format = '+#,##0"円";-#,##0"円";0"円"'
                        cell.font = green_font if v >= 0 else red_font
                        cell.alignment = right_align
                    except (ValueError, TypeError):
                        cell.value, cell.alignment = val, right_align
                elif col in pct_cols:
                    try:
                        v = float(val); cell.value = v / 100.0
                        cell.number_format = '+0.000%;-0.000%;0.000%'
                        cell.font = green_font if v >= 0 else red_font
                        cell.alignment = right_align
                    except (ValueError, TypeError):
                        cell.value, cell.alignment = val, right_align
                else:
                    cell.value = val
                    cell.alignment = left_align if isinstance(val, str) and len(str(val)) > 4 else center
                cell.border = border
        for ci, col in enumerate(cols, start=1):
            mx = max([len(str(col))] + [len(str(v)) for v in df[col].head(200).fillna("")])
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 32)
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "仮想日次サマリー"
    _write(ws1, df_hist, {"合計損益(円)"}, {"損益率(%)"})
    if not df_det.empty:
        ws2 = wb.create_sheet("仮想銘柄別明細")
        _write(ws2, df_det, {"損益(円)"}, {"OC騰落率(%)"})
    wb.save(PAPER_XLSX)
    print(f"  ✅ Excel保存: {PAPER_XLSX.name} ({len(df_hist)}日 / {len(df_det)}明細)")


def main():
    print("=" * 65)
    print(f"  仮想売買（ペーパートレード）損益記録  {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("=" * 65)

    pv = VERIFY_CAPITAL
    df_sig, date_fmt = _latest_signal()
    if df_sig is None:
        print("  ❌ シグナルファイルがありません")
        sys.exit(1)
    print(f"  シグナル日付(米国): {date_fmt}  運用資産(仮想): {pv:,.0f}円")

    try:
        token = ko.get_token()
    except Exception:
        try:
            token = ko.get_token(force_refresh=True)
        except Exception as e:
            print(f"  ❌ トークン取得失敗（kabuステーション未起動?）: {e}")
            ko._send_gmail("[投資戦略/仮想] ⚠️ 価格取得失敗",
                           f"kabuステーション未起動/未ログインの可能性。\n{e}")
            sys.exit(1)

    detail_rows = []
    total_pnl = 0.0
    for _, row in df_sig.iterrows():
        ticker = str(row["Ticker"])
        pos = float(row["ポジション"])
        if pos == 0:
            continue
        code = ko.JP_TICKER_TO_CODE.get(ticker, "")
        name = ko.JP_NAMES.get(ticker, ticker)
        direction = "★ LONG" if pos > 0 else "▼ SHORT"

        if pos < 0 and ticker in ko.SHORT_SKIP_TICKERS:
            print(f"    [{ticker}] {name} SHORT スキップ（デイトレ売建非対応）")
            continue
        if not code:
            continue

        op, cl = _board_open_close(token, code)
        if not op or not cl:
            detail_rows.append({"日付": date_fmt, "銘柄": ticker, "名称": name,
                                "方向": direction, "始値": op or "", "終値": cl or "",
                                "数量": "", "損益(円)": "", "OC騰落率(%)": "",
                                "備考": "価格取得不可"})
            print(f"    [{ticker}] {name} 価格取得不可 → スキップ")
            continue

        lot = ko.LOT_SIZE.get(code, 1)
        qty = int(pv * abs(pos) / op / lot) * lot
        if qty == 0:
            continue
        pnl = float(np.sign(pos)) * qty * (cl - op)
        oc = (cl - op) / op * 100.0
        total_pnl += pnl
        detail_rows.append({"日付": date_fmt, "銘柄": ticker, "名称": name,
                            "方向": direction, "始値": round(op, 1), "終値": round(cl, 1),
                            "数量": qty, "損益(円)": round(pnl), "OC騰落率(%)": round(oc, 3),
                            "備考": "仮想(板)"})
        print(f"    [{ticker}] {name} {direction} {qty}口  "
              f"始値{op:,.0f}→終値{cl:,.0f}  仮想損益 {pnl:+,.0f}円")

    traded = [r for r in detail_rows if r["備考"] == "仮想(板)"]
    pos_count = len(traded)
    pnl_pct = (total_pnl / pv * 100.0) if pv else 0.0
    win = "W" if total_pnl >= 0 else "L"

    summary_row = {"日付": date_fmt, "合計損益(円)": round(total_pnl),
                   "損益率(%)": round(pnl_pct, 3), "勝敗": win, "ポジション数": pos_count}

    _upsert_csv(PAPER_DETAIL_CSV, DETAIL_FIELDS, detail_rows, key_cols=["日付", "銘柄"])
    _upsert_csv(PAPER_HIST_CSV, HIST_FIELDS, [summary_row], key_cols=["日付"])
    print(f"\n  仮想合計損益: {total_pnl:+,.0f}円 ({pnl_pct:+.3f}%) {win}  / {pos_count}銘柄")
    _save_excel()

    # 通知
    body = [f"仮想売買 結果（{date_fmt}）",
            f"合計損益: {total_pnl:+,.0f}円 ({pnl_pct:+.3f}%)  {win}  / {pos_count}銘柄",
            ""]
    for r in traded:
        body.append(f"  [{r['銘柄']}] {r['名称']} {r['方向']} {r['数量']}口  "
                    f"{r['始値']}→{r['終値']}  {r['損益(円)']:+}円")
    ko._send_gmail(f"[投資戦略/仮想] {date_fmt} {total_pnl:+,.0f}円 {win}", "\n".join(body))
    print("\n=== 完了 ===")


if __name__ == "__main__":
    main()
